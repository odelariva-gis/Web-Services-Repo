'''
Utilies for serveral commonly used workflows for GraphQL Urban API. 
'''


import os
from typing import List
import sys
from arcgis.gis import GIS
import config
import importlib
import pandas as pd
import time
import config
from datetime import date
import requests
import openpyxl as pxl
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Callable
import pickle

importlib.reload(config)

class DidNotAppend(Exception):
    pass



def timer_decorator(file_path)->Callable:

    '''
    Decorator function to time other functions...
    '''

    def middle(in_function):

        def wrap_function(*args, **kwargs):
            t1= time.time()
            result = in_function(*args, **kwargs)
            t2 = time.time()
            time_delta = t2 - t1

            lil_spacer()
            lil_bugger()
            if time_delta < 60:
                print(f'Function {in_function.__name__!r} was executed in {(t2-t1):.4f}s!')
                with open(file_path, 'a') as f:
                    f.write(f"{in_function.__name__}: {(t2-t1):.4f} seconds \n")

            else:
                print(f'Function {in_function.__name__!r} was executed in {((t2-t1)/60):.4f}m!')
                with open(file_path, 'a') as f:
                    f.write(f"{in_function.__name__}: {((t2-t1)/60):.4f} minutes \n")

            lil_bugger()
            lil_spacer()

            
            return result
        return wrap_function
    return middle


def lil_bugger()-> None:

    print("*"*50)

    return None

def lil_spacer()-> None:

    print("\n")

    return None

def lil_dashy()-> None:

    print("-" * 50)

    return None

def return_today()-> date:
    '''
    Rreturn today's dates
    '''

    today = date.today()
    date_ = today.strftime("%Y%m%d")

    return date_

# def lil_bugger(in_function)-> Callable:
#     '''
#     Decorator function to include asterisks after function...
#     '''

#     bugs_ = "\n" + "*"*50

#     def wrap_function(*args, **kwargs):
#         return(f"{in_function} {bugs_}")
#     return wrap_function

# def lil_spacer(in_function)-> Callable:
#     '''
#     Decorator function to include space after function...
#     '''

#     space_ = "\n"
#     def wrap_function(*args, **kwargs):
#         return(f"{in_function} {space_}")
#     return wrap_function

# def lil_dashy(in_function)-> Callable:
#     '''
#     Decorator functiion to add dashes after function...
#     '''
    
#     dash_ = "\n" + "-" *50
    
#     def wrap_function(*args, **kwargs):
#         return (f"{in_function} {dash_}")
#     return wrap_function

@timer_decorator(f"timing_log_{return_today()}.txt")
def loggin_agol(config_file_name: str) -> GIS:

    '''
    Will log into AGOL for you! Must keep your 'config.py' file in the same location as your file. 

    Function return portal object of type GIS.  
    '''
    
    # Initialize/read config file
    cwd = sys.path[0]
    config_file = os.path.join(cwd, config_file_name)
    
    if not os.path.exists(config_file):
        print(f"Config file not found: {config_file_name}")
        sys.exit()
    else:
        print('Config File found and will continue!')

    # Getting login information
    username   = config.login_dict['username']
    pw         = config.login_dict['pw']
    portal_url = config.login_dict['portal_url']

    # Login to the portal...
    print(f'Loging in as {username} into {portal_url}! Please wait...')

    source = GIS(portal_url, username, pw)
    print(f'Success! Logged into {source} as {source.properties.user.username}!')

    return source



def create_token_header(config_file: str, gis_source: GIS = None) -> dict:

    '''
    This will function return the auth token string for GraphQL endpoint headers
    a type dict.  

    User can either provide a config.py file to input the token and have the function
    create the GraphQL header, or provide a GIS source of type GIS to provide the function. 
    This funciton will work in conjunction the function above, loggin_agol above, to provide a source
    of type GIS. 
    '''
    
    # Initialize/read config file
    cwd = sys.path[0]
    config_file = os.path.join(cwd, config_file)
    if not os.path.exists(config_file):
        print(f"Config file not found: {config_file}")
        sys.exit()
    else:
        print('Config File found and will continue!')

    token = ''

    if gis_source:
        print(f"Will get token from the source GIS here: {gis_source}")
        token = gis_source._con.token
    else:
        print('Will get token from configure file...')
        token = config_file.login_dict['token']

    print(f'Creating your endpoint headers with token...')
    print(" ")

    endpoint_header = {'Authorization': 'Bearer ' + token}

    return endpoint_header


def request_token(gis_source: GIS) -> str:

    '''
    Returns token for user as type string.
    '''
    token = ''

    if gis_source:
        print(f"Will get token from the source GIS here: {gis_source}")
        token = gis_source._con.token
    
    print('Here is your current token: ')
    lil_spacer()

    print(token)

    return token

@timer_decorator(f"timing_log_{return_today()}.txt")
def log_in_source(gis_source: GIS, token_: str)-> None:
    '''
    Logs into the source 
    '''

    # params = {'f': 'json', 'token': token_}
    source_users = gis_source.users.search(query='', sort_field='username', sort_order='asc', max_users=10000, outside_org=False, exclude_system=True)
    print('Logged in to {}'.format(gis_source))

    return source_users

@timer_decorator(f"timing_log_{return_today()}.txt")
def get_gis_content(gis_source)->List:
    '''
    Pulls GIS items of all owners and type Feature Service
    '''

    item_list = gis_source.content.search(query='owner:*', item_type = 'Feature Service', max_items = 1000)

    print("We have {} feature services to process.".format(len(item_list)))

    return item_list

def pop_empty_urls(item_list: list)-> list:
    '''
    Pop out empty URL
    '''

    new_item_list = []

    print('Popping out empty urls...')

    for item_ in item_list:
        if item_.url is None or item_.url == '':
            print(f"{item_} has an empty URL string and will be removed")
        else:
            new_item_list.append(item_)

    print('Finished running the script!')

    return new_item_list


def pop_gdb_urls(item_list: list)-> list:
    '''
    Pop out url ending with a gdb. 
    '''

    new_item_list = []

    for item_ in item_list:
        if item_.url[-4:] == '.gdb':
            print(f"{item_} is in a file geodatabse and will be removed.")
        else:
            new_item_list.append(item_)

    print("Finished cleaning URL's ending with GDB.")

    return new_item_list


def pop_repeated_urls(item_list: list) -> list:
    '''
    Pop out repeated URLs!
    '''

    non_repeated_urls = []
    new_item_list = []

    counter_ = 0

    for item_ in item_list:
        if item_.url not in non_repeated_urls:
            non_repeated_urls.append(item_.url)
            new_item_list.append(item_)
        else:
            counter_ += 1

    print(f"We found {counter_} repeated URLs and they have been removed.")
    print("Finished cleaning repeated URL's...")

    return new_item_list

def clean_urls(item_list: list) -> list:

    '''
    Cleans the URL!
    '''

    print("Clearing numeric URL's")

    for item_ in item_list:
        counter_list = []

        len_url = len(item_.url)
        counter_= 0
        if item_.url[len_url-1:].isdigit():
            for letter in item_.url:
                counter_ = counter_ + 1
                if letter == chr(47):
                    counter_list.append(counter_)
            slash_remover = counter_list[-1:][0]
            item_.url = item_.url[:slash_remover]
            print(f"Removed numberic endings from {item_.title}")
        
    print("Finished clearning numeric URL's")

    return item_list

def check_status_error(response_json, status_, error_)->str:
    '''
    Checking for status and error
    '''

    try:
        if response_json['status']:
            status_ = True
    except:
        pass

    try:
        if response_json['error']:
            error_ = True
    except:
        pass
        
    print(f"Status: {status_}")
    print(f"Error: {error_}")

    return status_, error_

@timer_decorator(f"timing_log_{return_today()}.txt")
def return_json(url_: str, params_: dict):
    '''
    returns json items
    '''
    response = requests.post(url_, params = params_, verify = False)

    return response.json()


@timer_decorator(f"timing_log_{return_today()}.txt")
def pull_json(item_list: List, params: dict) -> dict:
    '''
    iterate through items
    '''
    
    print("Starting iterations now...")
    lil_spacer()

    dict_ = {}
    url_dict_ = {}

    for item_ in item_list:
        
        print(f"Iterating through: {item_.title}")
        print(f"Inital URL: {item_.url}")
        try:
            error_      = False
            status_     = False
            all_errors_ = False
            
            current_url = item_.url.replace('rest','admin').replace('/MapServer','.MapServer') + '/iteminfo/manifest/manifest.json'
            print(f"Updated URL: {current_url}")
            
            ###
            # response = requests.post(current_url, params = params, verify = False)
            # response_json = response.json()
            response_json = return_json(current_url, params)
            status_, error_ = check_status_error(response_json, error_, status_)
            
            if status_ or error_:
                print("Found an error when pulling the JSON String MapServer to MapServer")
                all_errors_ = True
        
            if all_errors_:
                print("Did not append to dictionary!")
                lil_spacer()
                raise DidNotAppend
            else:
                url_dict_[item_.title] = current_url
                dict_[item_] = response_json
                print(f"Appended the following: {response_json}")  
                lil_spacer()

        ### except here    

        # except DidNotAppend:
        #     print("Did not append to dictionary...")
        #     print("JSON Response Failed when using /MapServer to .MapServer, will try another path... ")
        except:
            print("JSON Response Failed when using /MapServer to .MapServer, will try another path... " )
            
            error_      = False
            status_     = False
            all_errors_ = False
            
            try:
                current_url = item_.url.replace('rest','admin').replace('/FeatureServer','.MapServer') + '/iteminfo/manifest/manifest.json'
                print(f"Updated URL: {current_url}")

                response = requests.post(current_url, params = params, verify = False)
                response_json = response.json()
                status_, error_ = check_status_error(response_json, error_, status_)

                if status_ or error_:
                    print("Found error when pulling the JSON string FeatureServer to MapServer")
                    all_errors_ = True

                if all_errors_:
                    print("Did not append to dictionary!")
                    lil_spacer()
                    raise DidNotAppend
                else:
                    url_dict_[item_.title] = current_url
                    dict_[item_] = response_json
                    print(f"Appended the following to dictionary: ")
                    print(response_json)
                    lil_spacer()
            
        ### except here
            except DidNotAppend:
                print("Did not append to dictionary...")
                print("JSON Response Failed when using /MapServer to .MapServer, will try another path... ")
            
            except:
                print("JSON Response Failed when using /MapServer to .MapServer, will try another path... " ) 
                pass
        
        lil_dashy()
        lil_spacer()
        
    print("Finished iterating through all items!") 

    return dict_, url_dict_

@timer_decorator(f"timing_log_{return_today()}.txt")
def iterate_json(dict_:dict, url_dict_:dict)->List:
    '''
    iterates through json
    '''

    main_list = []
    service_counter = 0

    hosted_list = []

    print('Iterating through services...')
    for k, v in dict_.items():
        service_counter_ = 0
        list_            = []
        
        print(f"Processing {k.title} feature service...")
        
        print(k.title)
        print(k.owner)
        print(k.url)
        
        list_.append(k.title)
        list_.append(k.owner)
        list_.append(k.url)


        ### Hosted difference can go here TODO
        
        if 'rest/services/Hosted' in k.url:
            hosted_ = []
            hosted_.append(k.title)
            hosted_.append(k.owner)
            hosted_.append(k.url)
            hosted_db = v['databases'][0]['onPremiseConnectionString'].replace('DATABASE=', '')
            hosted_.append(hosted_db)
        
            hosted_list.append(hosted_)
            
            del hosted_
            
        else:
            try:
                list_.append(url_dict_[k.title])
            except:
                list_.append('Missing URL')

            lil_spacer()

            ### deal with idtem with no databases here TODO 

            on_ser_con = v['databases'][0]['onServerConnectionString'].split(';')
            on_ser_con_len = len(on_ser_con)

            print(f"Your onServerConnection String is length of: {on_ser_con_len}")

            if v['databases'][0]['onServerConnectionString'][-4:] == '.gdb':
                print("Response ends with a GDB, will not append onServer data...")

                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
        

            elif on_ser_con_len == 1:
                list_.append(on_ser_con)
                list_.append(" ")
                list_.append(" ")

            else:
                e_ = 0
                
                if on_ser_con_len == 10:
                    e_ = 1
                
                print("Reponse does not end with GDB, checking onServerConnection string...")
                split_ = v['databases'][0]['onServerConnectionString'].split(';')
                print(split_)

                if split_[2 + e_] is None or split_[2 + e_] == '':
                    list_.append("No onServerConnection Instance")
                    print('No onServerConnectionn Instance')
                else:
                    list_.append(split_[2 + e_].replace("INSTANCE=", ''))
                    print(f"Appended ON_SER_INSTANCE: {split_[2 + e_]}")

                if split_[3 + e_] is None or split_[3 + e_] == '':
                    list_.append("No onServerConnection DB Client")
                    print('No onServerConnection DB Client')
                else:
                    list_.append(split_[3 + e_].replace("DBCLIENT=", ''))
                    print(f"Appended ON_SER_DB_CLIENT: {split_[3 + e_]}")

                
                if split_[4 + e_] is None or split_[4 + e_] == '':
                    list_.append("No onServerConnection DB Connection")
                    print('No onServerConnection DB Connection')
                else:
                    list_.append(split_[4 + e_].replace("DB_CONNECTION_PROPERTIES=", ''))
                    print(f"Appended ON_SER_DB_CONNECT: {split_[4 + e_]}")

                if split_[5 + e_] is None or split_[5 + e_] == '':
                    list_.append("No onServerConnection Database")
                    print('No onServerConnection Database')
                else:
                    list_.append(split_[5 + e_].replace("DATABASE=", ''))
                    print(f"Appended ON_SER_DATABASE: {split_[5 + e_]}")
                    
                if split_[6 + e_] is None or split_[6 + e_] == '':
                    list_.append("No onServerConnection User")
                    print('No onServerConnection User')
                else:
                    list_.append(split_[6 + e_].replace("USER=", ''))
                    print(f"Appended ON_SER_USER: {split_[6 + e_]}")
                    
                if split_[7 + e_][:4] == 'AUTH':
                    list_.append(split_[7 + e_].replace("AUTHENTICATION_MODE=", '').replace("AUTHENTICATION_MODE=",''))
                    print(f"Appended ON_SER_AUTH: {split_[7 + e_]}")

                    list_.append(split_[8 + e_].replace("VERSION=", '').replace("BRANCH=", ''))
                    print(f"Appended ON_SER_VERSION: {split_[8 + e_]}")

                elif split_[7 + e_][:4] == 'BRAN' or split_[7 + e_][:4] == 'VERS':
                    list_.append(split_[8 + e_].replace("VERSION=", '').replace("BRANCH=", ''))
                    print(f"Appended ON_SER_VERSION: {split_[8 + e_]}")

                    list_.append(split_[7 + e_].replace("AUTHENTICATION_MODE=", '').replace("AUTHENTICATION_MODE=",''))
                    print(f"Appended ON_SER_AUTH: {split_[7 + e_]}")
                else:
                    list_.append(' ')
                    list_.append(' ')
                    
                del split_

            lil_spacer()

            on_prem_con = v['databases'][0]['onPremiseConnectionString'].split(';')
            on_prem_con_len = len(on_prem_con)

            print(f"Your onPremiseConnection String is length of: {on_prem_con_len}")
            print(on_prem_con)

            if v['databases'][0]['onPremiseConnectionString'][-4:] == '.gdb':
                print("Response ends with a GDB, will not append onPremise data...")

                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
                list_.append('Service ends in GDB')
            
                lil_spacer()

            elif on_prem_con_len == 1:
                list_.append(on_prem_con)
                list_.append(" ")
                list_.append(" ")

            else:
                e_ = 0
                if on_prem_con_len == 10:
                    e_ = 1
                
                print("Reponse does not end with GDB, checking onPremiseConnection string...")
                split_ = v['databases'][0]['onPremiseConnectionString'].split(';')

                if split_[2 + e_] is None or split_[2 + e_] == '':
                    list_.append("No onPremiseConnection Instance")
                    print('No onPremiseConnection Instance')
                else:
                    list_.append(split_[2 + e_].replace("INSTANCE=", ''))
                    print(f"Appended ON_PREM_INSTANCE: {split_[2 + e_]}")

                if split_[3 + e_] is None or split_[3 + e_] == '':
                    list_.append("No onPremiseConnection DB Client")
                    print('No onPremiseConnection DB Client')
                else:
                    list_.append(split_[3 + e_].replace("DBCLIENT=", ''))
                    print(f"Appended ON_PREM_DB_CLIENT: {split_[3 + e_]}")

                
                if split_[4 + e_] is None or split_[4 + e_] == '':
                    list_.append("No onPremiseConnection DB Connection")
                    print('No onPremiseConnection DB Connection')
                else:
                    list_.append(split_[4 + e_].replace("DB_CONNECTION_PROPERTIES=", ''))
                    print(f"Appended ON_PREM_DB_CONNECT: {split_[4 + e_]}")

                if split_[5 + e_] is None or split_[5 + e_] == '':
                    list_.append("No onPremiseConnection Database")
                    print('No onPremiseConnection Database')
                else:
                    list_.append(split_[5 + e_].replace("DATABASE=", '').replace('PROJECT_INSTANCE=', ''))
                    print(f"Appended ON_PREM_DATBASE: {split_[5 + e_]}")
                    
                if split_[6 + e_] is None or split_[6 + e_] == '':
                    list_.append("No onPremiseConnection User")
                    print('No onPremiseConnection User')
                else:
                    list_.append(split_[6 + e_].replace("USER=", ''))
                    print(f"Appended ON_PREM_USER: {split_[6 + e_]}")
                    
                if split_[7 + e_][:4] == 'AUTH':
                    list_.append(split_[7 + e_].replace("AUTHENTICATION_MODE=", ''))
                    print(f"Appended ON_PREM_AUTH: {split_[7 + e_]}")
                    
                    list_.append(split_[8 + e_].replace("VERSION=", '').replace("BRANCH=", ''))
                    print(f"Appended ON_PREM_VERSION: {split_[8 + e_]}")
                    
                elif split_[7 + e_][:4] == 'BRAN' or split_[7 + e_][:4] == 'VERS':
                    list_.append(split_[8 + e_].replace("VERSION=", '').replace("BRANCH=", ''))
                    print(f"Appended ON_PREM_VERSION: {split_[8 + e_]}")

                    list_.append(split_[7 + e_].replace("AUTHENTICATION_MODE=", ''))
                    print(f"Appended ON_PREM_AUTH: {split_[7 + e_]}")
                else:
                    list_.append(' ')
                    list_.append(' ')
                
                del split_
                lil_spacer()      

            print(f"Getting the number of the {k.title} databases...")

            db_len = len(v['databases'])
            print(f"{k.title} has {db_len} of databases...")
            for i in range(db_len):
                #print(v['databases'][i]['datasets'])
                ds_len = len(v['databases'][i]['datasets'])
                for j in range(ds_len):
                    service_ = v['databases'][i]['datasets'][j]['onServerName']
                    print(f"On ServerName/Services: {service_}")
                    list_.append(service_)
                    service_counter_ = service_counter_ + 1

            main_list.append(list_)
            # hosted_list.append(hosted_)

            if service_counter_ > service_counter:
                service_counter = service_counter_

            del list_, service_counter_

        lil_spacer()
        lil_dashy()
    
    print("Tool has finished running!")

    return main_list, hosted_list, service_counter
   

def output_to_excel(path_:str, output_df: pd.DataFrame , hosted_df: pd.DataFrame ) -> None:
    '''
    outputs to excel
    '''  

    if os.path.exists(path_):
        os.remove(path_)
        print("{0} has been deleted.".format(path_))

    output_df.to_excel(path_, index=False, sheet_name = "Services")
    excel_book = pxl.load_workbook(path_)
    excel_book.create_sheet('Hosted')
    rows = dataframe_to_rows(hosted_df, index = False)

    ws = excel_book["Hosted"]

    for r_idx, row in enumerate(rows, 1):
    #print(r_idx)
        for c_idx, value in enumerate(row, 1):
        #  print(c_idx)
            ws.cell(row = r_idx, column = c_idx, value = value)

    excel_book.save(path_)

    print(f'Excel file successfuly created at {path_}!') 

    return None   

def output_to_excel_xy(path_:str, output_df: pd.DataFrame ) -> None:
    '''
    outputs to excel for xy coordinates
    '''  

    # columns_to_pd = ['TITLE', 'XCOORD', 'YCOORD']

    # ### Setting up extended field names


    # ### Setting up field names for extended columns
    # output_df = pd.DataFrame(list_, columns = columns_to_pd)

    if os.path.exists(path_):
        os.remove(path_)
    print("{0} has been deleted.".format(path_))

    output_df.to_excel(path_, index=False)


    print(f'Excel file successfuly created at {path_}!') 

    return None   

def return_xy_dict(item_list: List, params_: dict)-> dict: 
    '''
    returns the xmin and xmax of a service
    '''
    title_dict = {}
    for item in item_list:
        title_ = item.title 
        url_ = item.url
        print(f"Processing {title_}!")
        json_ = return_json(url_, params_)

        try:
            xmin = json_['initialExtent']['xmin']
            ymin = json_['initialExtent']['ymin']
            xmax = json_['initialExtent']['xmax']
            ymax = json_['initialExtent']['ymax']
            sr_  = json_['spatialReference']['wkid']
            coord_dict = {'xmin': xmin, 'ymin': ymin, 'xmax': xmax, 'ymax': ymax, 'sr': sr_}
            title_dict[title_] = coord_dict
        except:
            print('Initial extent not found.')
              
    return title_dict

def return_xy_list(xy_dict: dict) -> List:
    title_xy = []
    for k,v in xy_dict.items():
        temp_list = []
        print(f"Item: {k}")
        x_mid = (v['xmax'] + v['xmin'])/2
        y_mid = (v['ymax'] + v['ymin'])/2
        
        sr_ = v['sr']
        
        print(x_mid)
        print(y_mid)
        
        temp_list.append(k)
        temp_list.append(x_mid)
        temp_list.append(y_mid)
        temp_list.append(sr_)
        title_xy.append(temp_list)
        
        del temp_list, x_mid, y_mid, 
        S
        lil_spacer()
        
    return title_xy

def pickle_this(url_dict: dict, name_)-> None:
    '''
    Pickles the dictionary if needed
    '''

    with open(f'{name_}.pickle', 'wb') as handle:
        pickle.dump(url_dict, handle, protocol=pickle.HIGHEST_PROTOCOL)

    return None

def unpickle_that(pickle_loc: str)-> dict: 
    '''
    unpickles this
    '''
    with open(pickle_loc, 'rb') as handle:
        url_dict = pickle.load(handle)

    return url_dict
                
        
        





